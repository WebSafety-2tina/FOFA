"""
数据处理工具类
"""
import json
import re
from typing import List, Dict, Set, Optional
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from main.config import FofaConfig, ProxyConfig
from models.table_bean import TableBean, ExcelBean, TabDataBean


class DataUtil:
    """数据工具类"""
    
    # 端口正则模式
    PORT_PATTERN_443 = re.compile(r":443$")
    PORT_PATTERN_80 = re.compile(r":80$")
    
    @staticmethod
    def getValueFromIP(ip: str) -> float:
        """
        将IP地址转换为浮点数（用于排序）
        
        Args:
            ip: IP地址
            
        Returns:
            浮点数值
        """
        parts = ip.split('.')
        if len(parts) != 4:
            return 0.0
        try:
            return (float(parts[0]) * 1000000 + 
                   float(parts[1]) * 1000 + 
                   float(parts[2]) + 
                   float(parts[3]) * 0.001)
        except ValueError:
            return 0.0
    
    @staticmethod
    def replaceString(tabTitle: str) -> str:
        """
        替换Tab标题中的(*)标记
        
        Args:
            tabTitle: Tab标题
            
        Returns:
            处理后的查询字符串
        """
        if tabTitle.startswith("(*)"):
            tabTitle = tabTitle[3:]
            return f"({tabTitle}) && (is_honeypot=false && is_fraud=false)"
        return tabTitle
    
    @staticmethod
    def loadJsonData(
        bean: Optional[TabDataBean],
        obj: Dict,
        excelData: Optional[List[ExcelBean]],
        urlList: Optional[Set[str]],
        isExport: bool = False
    ) -> List:
        """
        加载JSON数据
        
        Args:
            bean: Tab数据Bean
            obj: JSON对象
            excelData: Excel数据列表（导出时使用）
            urlList: URL列表（导出时使用）
            isExport: 是否为导出模式
            
        Returns:
            数据列表
        """
        results = obj.get("results", [])
        if not results:
            return []
        
        config = FofaConfig.getInstance()
        fields = config.additionalField
        list_data = []
        
        for index, result_item in enumerate(results):
            if len(result_item) < 8:
                continue
            
            host = result_item[0] or ""
            title = result_item[1] or ""
            ip = result_item[2] or ""
            domain = result_item[3] or ""
            try:
                port = int(result_item[4]) if result_item[4] else 0
            except (ValueError, TypeError):
                port = 0
            protocol = result_item[5] or ""
            
            # 过滤非标准端口的http/https
            if port not in [443, 80] and protocol in ["http", "https"]:
                if not host.endswith(f":{port}"):
                    continue
            
            server = result_item[6] or ""
            link = result_item[7] or ""
            
            # 获取额外字段
            extra_fields = {
                "fid": "", "os": "", "icp": "", "product": "",
                "certs_subject_cn": "", "certs_subject_org": "", "lastupdatetime": ""
            }
            
            for i, field_name in enumerate(fields):
                if 8 + i < len(result_item):
                    if field_name in extra_fields:
                        extra_fields[field_name] = result_item[8 + i] or ""
            
            if isExport:
                # 导出模式
                data = ExcelBean(
                    host=host, title=title, ip=ip, domain=domain,
                    port=port, protocol=protocol, server=server
                )
                data.lastupdatetime = extra_fields["lastupdatetime"]
                data.fid = extra_fields["fid"]
                data.os = extra_fields["os"]
                data.product = extra_fields["product"]
                data.icp = extra_fields["icp"]
                data.certs_subject_cn = extra_fields["certs_subject_cn"]
                data.certs_subject_org = extra_fields["certs_subject_org"]
                
                # 去重处理（修复逻辑错误）
                if data in excelData:
                    try:
                        existing = excelData[excelData.index(data)]
                        if port in [443, 80]:
                            if ":443" in existing.host or ":80" in existing.host:
                                excelData.remove(existing)
                            elif ":443" in data.host or ":80" in data.host:
                                continue
                        if existing.host == data.host:
                            if existing.title:
                                continue
                            else:
                                excelData.remove(existing)
                    except (ValueError, IndexError):
                        # 如果找不到或已删除，继续添加
                        pass
                
                excelData.append(data)
                if link:
                    urlList.add(link)
            else:
                # 表格显示模式
                data = TableBean(
                    num=0, host=host, title=title, ip=ip, domain=domain,
                    port=port, protocol=protocol, server=server
                )
                data.fid = extra_fields["fid"]
                data.icp = extra_fields["icp"]
                data.os = extra_fields["os"]
                data.certCN = extra_fields["certs_subject_cn"]
                data.product = extra_fields["product"]
                data.certOrg = extra_fields["certs_subject_org"]
                data.lastupdatetime = extra_fields["lastupdatetime"]
                
                # 去重处理（修复逻辑错误）
                if data in list_data:
                    try:
                        existing = list_data[list_data.index(data)]
                        if port in [443, 80]:
                            if ":443" in existing.host or ":80" in existing.host:
                                data.num = existing.num
                                list_data.remove(existing)
                            elif ":443" in data.host or ":80" in data.host:
                                continue
                        if existing.host == data.host:
                            if existing.title:
                                continue
                            else:
                                data.num = existing.num
                                list_data.remove(existing)
                    except (ValueError, IndexError):
                        # 如果找不到或已删除，继续添加
                        pass
                
                if bean and data.num == 0:
                    bean.count += 1
                    data.num = bean.count
                
                if link and bean:
                    bean.dataList.add(link)
                
                list_data.append(data)
        
        return list_data
    
    @staticmethod
    def exportToExcel(
        fileName: str,
        tabTitle: str,
        totalData: List[ExcelBean],
        urls: List[List[str]],
        errorPage: str
    ):
        """
        导出数据到Excel
        
        Args:
            fileName: 文件名
            tabTitle: Tab标题
            totalData: 总数据列表
            urls: URL列表
            errorPage: 错误页面信息
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "查询结果"
            
            # 设置表头
            headers = ["HOST", "标题", "域名", "IP", "端口", "协议", "server指纹"]
            config = FofaConfig.getInstance()
            
            if "lastupdatetime" in config.additionalField:
                headers.append("最近更新时间")
            if "os" in config.additionalField:
                headers.append("操作系统")
            if "icp" in config.additionalField:
                headers.append("ICP")
            if "product" in config.additionalField:
                headers.append("产品指纹")
            if "fid" in config.additionalField:
                headers.append("fid")
            if "certs_subject_cn" in config.additionalField:
                headers.append("证书域名")
            if "certs_subject_org" in config.additionalField:
                headers.append("证书持有者组织")
            
            # 写入表头
            ws.append(headers)
            
            # 设置表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, size=14, color="FFFFFF")
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 写入数据
            for data in totalData:
                row = [
                    data.host, data.title, data.domain, data.ip, data.port,
                    data.protocol, data.server
                ]
                if "lastupdatetime" in config.additionalField:
                    row.append(data.lastupdatetime)
                if "os" in config.additionalField:
                    row.append(data.os)
                if "icp" in config.additionalField:
                    row.append(data.icp)
                if "product" in config.additionalField:
                    row.append(data.product)
                if "fid" in config.additionalField:
                    row.append(data.fid)
                if "certs_subject_cn" in config.additionalField:
                    row.append(data.certs_subject_cn)
                if "certs_subject_org" in config.additionalField:
                    row.append(data.certs_subject_org)
                ws.append(row)
            
            # 设置列宽
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 38
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            
            # 创建URLs工作表
            ws2 = wb.create_sheet("urls")
            for url in urls:
                ws2.append(url)
            ws2.column_dimensions['A'].width = 40
            
            # 保存文件
            wb.save(fileName)
            
            message = f"导出成功！文件保存在 {fileName}"
            if errorPage:
                message = f"部分数据导出成功，其中第{errorPage}页加载失败，文件保存在 {fileName}"
            
            return True, message
        except Exception as e:
            return False, f"导出失败！{str(e)}"
    
    @staticmethod
    def loadConfigure() -> FofaConfig:
        """
        从配置文件加载配置
        
        Returns:
            FofaConfig实例
        """
        config = FofaConfig.getInstance()
        proxyConfig = ProxyConfig.getInstance()
        
        # 获取配置文件路径
        config_path = Path(__file__).parent.parent / "config.properties"
        
        if not config_path.exists():
            # 如果配置文件不存在，使用默认值
            return config
        
        try:
            # Properties文件格式：key=value，直接读取文件
            with open(config_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    # 跳过注释和空行
                    if not line or line.startswith('#'):
                        continue
                    
                    # 解析key=value格式
                    if '=' in line:
                        key, value = line.split('=', 1)
                        key = key.strip()
                        value = value.strip()
                        
                        if key == 'api':
                            config.API = value
                        elif key == 'key':
                            config.setKey(value)
                        elif key == 'max_size' or key == 'maxSize':
                            config.setSize(value)
                        elif key == 'check_status' or key == 'checkStatus':
                            config.checkStatus = value.lower() == 'on'
                        elif key == 'proxy_status' or key == 'proxyStatus':
                            proxyConfig.status = value.lower() == 'on'
                        elif key == 'proxy_type' or key == 'proxyType':
                            proxyConfig.proxy_type = value
                        elif key == 'proxy_ip' or key == 'proxyIp':
                            proxyConfig.proxy_ip = value
                        elif key == 'proxy_port' or key == 'proxyPort':
                            proxyConfig.proxy_port = value
                        elif key == 'proxy_user' or key == 'proxyUser':
                            proxyConfig.proxy_user = value
                        elif key == 'proxy_password' or key == 'proxyPassword':
                            proxyConfig.proxy_password = value
            
        except FileNotFoundError:
            print(f"配置文件不存在: {config_path}")
        except Exception as e:
            print(f"加载配置失败: {e}")
        
        return config

